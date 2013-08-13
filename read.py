from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import xml.etree.ElementTree as ET
import os
from docx import *
from ldModule import *
import re
import signal


def auto_match1(str1, str2):
    pass

def clean_parse(s):
    s = s.upper()
    s = re.sub('[^0-9A-Z]+', '', s)
    return s

def auto_match(str1, str2):
    location = find_match(str1, str2)
    
    matched = str1[location:location+len(str2)]
    ld = levenshtein(matched, str2)
    print "ld = %s" % ld
    print "length = %s" % len(matched)
    return matched, ld

def main():
    # obtin the current dirtory
    #current_dir = os.path.dirname(__file__)


    # obtain the current directory name
    # __file__ is the inner representation of this file
    current_dir = os.path.dirname(os.path.abspath(__file__))
    # join the path to go into the subfolder
    cv_dir = os.path.join(current_dir, 'Bi/CVs/')

    work_path = os.path.join(current_dir, 'wos_new.xlsx')
    result_path = os.path.join(current_dir, 'new.xlsx')
    
    work_book = load_workbook(work_path)
    sheet = work_book.worksheets[0]

    result_book = load_workbook(result_path)
    result_sheet = result_book.worksheets[0]
    result_rows = result_sheet.rows
    count = 1
    
    tmp_name = os.walk(cv_dir)

    for root, sub_dir, files_list in tmp_name:
        for file in files_list:
            if file[-4:] == 'docx':
                
                file = os.path.join(root,file)

                # the root is in format of /blah/blah/CVs/10007
                # and cv_id will be '10007'
                cv_id =  root.rsplit("/",1)[1]
                added_cv_id = str(int(cv_id) + 20000)

                # TODO: add exception handler and log file support
                document = opendocx(file)
                unparsed_text = getdocumenttext(document)
                parsed_text = ''
                for c in unparsed_text:
                    parsed_text += c.encode("utf-8")
                    
                parsed_cv_text = clean_parse(parsed_text)
                
                        
                #string_len = len(string_text)
                cv_len = len(parsed_cv_text)
                #print "string length %d" % string_len
                #print "cv length %d" % cv_len
                author_row = sheet.rows
                
                #print len(author_row)
                for i in range(len(author_row)):
                    wos_id = str(author_row[i][6].value)
                    if wos_id == added_cv_id:

                        # title match
                        title_string = str(author_row[i][7].value)
                        title_string = clean_parse(title_string)
                        
                        wos_len = len(title_string)
                        #print "before auto_match"
                        #print parsed_cv_text
                        #print title_string
                        result = auto_match(parsed_cv_text, title_string)
                        ld = result[1]
                        match_string = result[0]
                        
                        #print "author %s, cv %s" % (str(author_row[i][6].value), cv_id)
                        #print ld
                        
                        result_sheet.cell(row = count, column = 0).value = added_cv_id
                        result_sheet.cell(row = count, column = 1).value = str(author_row[i][3].value)
                        result_sheet.cell(row = count, column = 2).value = 'cv-title'
                       
                        match_len = len(match_string)
                        print "matched len = %s" % match_len
                        prob = float(ld)/match_len
                        print prob
                        result_sheet.cell(row = count, column = 3).value = prob
                        if prob != 0:
                            result_sheet.cell(row = count, column = 4).value = 'OK'
                        else:
                            result_sheet.cell(row = count, column = 4).value = 'Prob is zero'
                        result_sheet.cell(row = count, column = 5).value = match_string
                        result_sheet.cell(row = count, column = 6).value = title_string
                        count += 1

                        # co-author match
                        if len(author_row[i]) >= 9:
                            co_author_str = str(author_row[i][2].value)+str(author_row[i][9].value)
                        else:
                            co_author_str = str(author_row[i][2].value)
                        author_list = co_author_str.split(';')
                        for co_author in author_list:
                            co_author = clean_parse(co_author)
                            result = auto_match(parsed_cv_text, co_author)
                            ld = result[1]
                            match_string = result[0]
                            #print 'co-author'+str(ld)
                            author_len = len(co_author)
                            prob = float(ld)/author_len
                            result_sheet.cell(row = count, column = 0).value = added_cv_id
                            result_sheet.cell(row = count, column = 1).value = str(author_row[i][3].value)
                            result_sheet.cell(row = count, column = 2).value = 'cv-co_author'
                            result_sheet.cell(row = count, column = 3).value = prob
                            if prob != 0:
                                result_sheet.cell(row = count, column = 4).value = 'OK'
                            else:
                                result_sheet.cell(row = count, column = 4).value = 'Prob is zero'
                            result_sheet.cell(row = count, column = 5).value = match_string
                            result_sheet.cell(row = count, column = 6).value = co_author
                            count += 1

                        # institution match
                        institution = str(author_row[i][4].value)
                        insti_list = institution.split(';')
                        for institution in insti_list:
                            institution = clean_parse(institution)
                            insti_len = len(institution)
                            result = auto_match(parsed_cv_text, institution)
                            ld = result[1]
                            match_string = result[0]
                            prob = float(ld)/insti_len
                            result_sheet.cell(row = count, column = 0).value = added_cv_id
                            result_sheet.cell(row = count, column = 1).value = str(author_row[i][3].value)
                            result_sheet.cell(row = count, column = 2).value = 'cv-institution'
                            result_sheet.cell(row = count, column = 3).value = prob
                            if prob != 0:
                                result_sheet.cell(row = count, column = 4).value = 'OK'
                            else:
                                result_sheet.cell(row = count, column = 4).value = 'Prob is zero'
                            result_sheet.cell(row = count, column = 5).value = match_string
                            result_sheet.cell(row = count, column = 6).value = institution
                            count += 1
                        
                        updated_result = ExcelWriter(workbook = result_book)
                        updated_result.save(filename = 'new.xlsx')
                        print 'another one'
                        
                        
'''
                for row in sheet.range('H2:H100'):
                    for cell in row:
                        string_text = str(cell.value)
                        
                        #if cv_id 
                
                        ld = levenshtein(parsed_cv_text, string_text)
                        print "ld length %d" % ld
                        print (ld-(cv_len-string_len))/float(string_len)
                        exit(0)
'''


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        exit(0)



'''
#print os.path.dirname(test)
#for s in os.listdir(test):
#    print s 
docu = opendocx(abs_dir)
raw_string = (getdocumenttext(docu))
str = ''
for c in raw_string:
str += c.encode('utf-8')
#print str
print '*'*80

article = 'article_eg.xlsx'
book = load_workbook(article)
sheet = book.worksheets[0]
for row in sheet.range('H2:H3'):
for cell in row:
print (cell.value)


print book.get_sheet_names()



book.save(filename = article)
'''